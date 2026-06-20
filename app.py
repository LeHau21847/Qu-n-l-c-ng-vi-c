import os, sys, json, mimetypes
import webbrowser
from threading import Timer
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
from flask import Flask, render_template, request, redirect, flash, url_for, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, text
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import qrcode, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from unidecode import unidecode

def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

template_dir = os.path.join(getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__))), 'templates')
static_dir = os.path.join(getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__))), 'static')

app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)
app.secret_key = os.environ.get('SECRET_KEY', 'cskv_super_secret_key_dev_2024')

uploads_dir = os.path.join(get_base_path(), 'uploads')
try:
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)
except Exception:
    pass
app.config['UPLOAD_FOLDER'] = uploads_dir
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB limit

# Database: dùng DATABASE_URL (PostgreSQL trên cloud) hoặc SQLite (local)
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Render dùng postgres://, SQLAlchemy cần postgresql://
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    db_path = os.path.join(get_base_path(), 'data_hoso.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + db_path.replace('\\', '/')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# --- BẢNG: Danh mục Loại Hồ sơ ---
class LoaiHoSo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ten_loai = db.Column(db.String(150), unique=True, nullable=False)
    cac_buoc = db.relationship('BuocXuLy', backref='loai_ho_so', lazy=True, cascade="all, delete-orphan", order_by='BuocXuLy.so_thu_tu')
    ho_so_list = db.relationship('HoSo', backref='loai_ho_so', lazy=True)

# --- BẢNG: Các bước cấu hình riêng cho từng Loại ---
class BuocXuLy(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loai_ho_so_id = db.Column(db.Integer, db.ForeignKey('loai_ho_so.id'), nullable=False)
    so_thu_tu = db.Column(db.Integer, nullable=False)
    ten_buoc = db.Column(db.String(150), nullable=False)

# --- BẢNG: Hồ sơ thực tế ---
class HoSo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ma_ho_so = db.Column(db.String(50), nullable=False)
    ten_nguoi_dan = db.Column(db.String(100), nullable=False)
    ten_khong_dau = db.Column(db.String(100), nullable=True) # --- CHO TÌM KIẾM ---
    loai_ho_so_id = db.Column(db.Integer, db.ForeignKey('loai_ho_so.id'), nullable=False)
    buoc_hien_tai_id = db.Column(db.Integer, db.ForeignKey('buoc_xu_ly.id'), nullable=True) 
    ngay_tao = db.Column(db.DateTime, default=datetime.now)
    ngay_cap_nhat_buoc = db.Column(db.DateTime, default=datetime.now)
    han_xu_ly = db.Column(db.DateTime, nullable=True)
    ghi_chu = db.Column(db.Text, nullable=True)
    is_tra_lai = db.Column(db.Boolean, default=False)
    ngay_sinh = db.Column(db.String(20), nullable=True)
    dia_chi = db.Column(db.String(255), nullable=True)
    nhan_khau_kem_theo = db.Column(db.Text, nullable=True)
    ten_chu_ho = db.Column(db.String(100), nullable=True)
    cccd_chu_ho = db.Column(db.String(50), nullable=True)
    quan_he_chu_ho = db.Column(db.String(50), nullable=True)
    cac_dinh_kem = db.relationship('HoSoDinhKem', backref='ho_so', lazy=True, cascade="all, delete-orphan")

    @property
    def so_ngay_ton(self):
        ref_date = self.ngay_cap_nhat_buoc or self.ngay_tao
        return (datetime.now() - ref_date).days

    @property
    def buoc_hien_tai(self):
        if self.buoc_hien_tai_id:
            return BuocXuLy.query.get(self.buoc_hien_tai_id)
        return None

    def is_hoan_tat(self):
        if not self.buoc_hien_tai_id:
            return False
        next_step = BuocXuLy.query.filter(
            BuocXuLy.loai_ho_so_id == self.loai_ho_so_id,
            BuocXuLy.so_thu_tu > self.buoc_hien_tai.so_thu_tu
        ).first()
        return next_step is None

    def get_trang_thai_label(self):
        if self.is_tra_lai:
            return "HỒ SƠ BỊ TRẢ LẠI"
        if not self.buoc_hien_tai_id:
            return "Chưa cấu hình bước"
        
        if self.is_hoan_tat():
            return f"Hoàn tất ({self.buoc_hien_tai.ten_buoc})"
        return self.buoc_hien_tai.ten_buoc

    def get_badge_class(self):
        if self.is_tra_lai:
            return "badge-danger"
        if not self.buoc_hien_tai_id:
            return "badge-secondary"
        if self.is_hoan_tat():
            return "badge-success-custom"
        
        if self.buoc_hien_tai.so_thu_tu == 1:
            return "badge-warning-custom"
        return "badge-custom"

# --- BẢNG: Sổ tay ghi chú (Quick Notes) ---
class QuickNote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    noi_dung = db.Column(db.Text, default="")
    ngay_cap_nhat = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

# --- BẢNG: Đính kèm tài liệu ---
class HoSoDinhKem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ho_so_id = db.Column(db.Integer, db.ForeignKey('ho_so.id'), nullable=False)
    file_name = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    file_data = db.Column(db.LargeBinary, nullable=True)  # Lưu file vào DB cho cloud
    ngay_tai = db.Column(db.DateTime, default=datetime.now)

# --- BẢNG: Tài khoản ---
class TaiKhoan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='user')
    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)
    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)

# --- BẢNG: Nhật ký hoạt động ---
class NhatKyHoatDong(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), nullable=False)
    hanh_dong = db.Column(db.String(255), nullable=False)
    chi_tiet = db.Column(db.Text, nullable=True)
    thoi_gian = db.Column(db.DateTime, default=datetime.now)

def tao_chuoi_khong_dau(text):
    if not text:
        return ""
    return unidecode(text).upper()

def ghi_nhat_ky(hanh_dong, chi_tiet=""):
    try:
        log = NhatKyHoatDong(username=session.get('username','system'), hanh_dong=hanh_dong, chi_tiet=chi_tiet)
        db.session.add(log)
        db.session.commit()
    except: db.session.rollback()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Bạn không có quyền truy cập!', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

with app.app_context():
    db.create_all()
    try:
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN ngay_cap_nhat_buoc DATETIME;"))
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN ten_khong_dau VARCHAR(100);"))
        db.session.commit()
    except: db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN han_xu_ly DATETIME;"))
        db.session.commit()
    except: db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN ngay_sinh VARCHAR(20);"))
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN dia_chi VARCHAR(255);"))
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN nhan_khau_kem_theo TEXT;"))
        db.session.commit()
    except: db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN ten_chu_ho VARCHAR(100);"))
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN cccd_chu_ho VARCHAR(50);"))
        db.session.execute(text("ALTER TABLE ho_so ADD COLUMN quan_he_chu_ho VARCHAR(50);"))
        db.session.commit()
    except: db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE ho_so_dinh_kem ADD COLUMN file_data BLOB;"))
        db.session.commit()
    except: db.session.rollback()

    if LoaiHoSo.query.count() == 0:
        loai_mk = LoaiHoSo(ten_loai="Khai báo nhân khẩu")
        db.session.add(loai_mk)
        db.session.commit()
        db.session.add(BuocXuLy(loai_ho_so_id=loai_mk.id, so_thu_tu=1, ten_buoc="Tiếp nhận đơn"))
        db.session.add(BuocXuLy(loai_ho_so_id=loai_mk.id, so_thu_tu=2, ten_buoc="Xác minh địa bàn"))
        db.session.add(BuocXuLy(loai_ho_so_id=loai_mk.id, so_thu_tu=3, ten_buoc="Trả kết quả"))
        db.session.commit()
        loai_tv = LoaiHoSo(ten_loai="Khai báo tạm vắng")
        db.session.add(loai_tv)
        db.session.commit()
        db.session.add(BuocXuLy(loai_ho_so_id=loai_tv.id, so_thu_tu=1, ten_buoc="Tiếp nhận/Ghi sổ"))
        db.session.add(BuocXuLy(loai_ho_so_id=loai_tv.id, so_thu_tu=2, ten_buoc="Ký duyệt chỉ huy"))
        db.session.commit()
    if QuickNote.query.count() == 0:
        db.session.add(QuickNote(noi_dung="Đây là sổ tay ghi chú của bạn..."))
        db.session.commit()
    if TaiKhoan.query.count() == 0:
        admin = TaiKhoan(username='admin', role='admin')
        admin.set_password('admin123')
        user = TaiKhoan(username='nhanvien', role='user')
        user.set_password('nhanvien123')
        db.session.add_all([admin, user])
        db.session.commit()

@app.context_processor
def inject_user():
    return dict(current_user=session.get('username'), current_role=session.get('role'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = TaiKhoan.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            ghi_nhat_ky('Đăng nhập hệ thống')
            return redirect(url_for('index'))
        flash('Sai tài khoản hoặc mật khẩu!', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    ghi_nhat_ky('Đăng xuất')
    session.clear()
    return redirect(url_for('login'))

# --- ROUTES DASHBOARD ---
@app.route('/')
@login_required
def index():
    tu_khoa = request.args.get('tu_khoa', '').strip()
    loai_id = request.args.get('loai_id', '').strip()
    tu_ngay = request.args.get('tu_ngay', '').strip()
    den_ngay = request.args.get('den_ngay', '').strip()

    query = HoSo.query

    if tu_khoa:
        tu_khoa_kd = tao_chuoi_khong_dau(tu_khoa)
        query = query.filter(or_(
            HoSo.ma_ho_so.ilike(f'%{tu_khoa}%'), 
            HoSo.ten_nguoi_dan.ilike(f'%{tu_khoa}%'),
            HoSo.ten_khong_dau.ilike(f'%{tu_khoa_kd}%')
        ))
    if loai_id and loai_id.isdigit():
        query = query.filter(HoSo.loai_ho_so_id == int(loai_id))
    try:
        if tu_ngay:
            dt_tu = datetime.strptime(tu_ngay, '%Y-%m-%d')
            query = query.filter(HoSo.ngay_tao >= dt_tu)
        if den_ngay:
            from datetime import timedelta
            dt_den = datetime.strptime(den_ngay, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(HoSo.ngay_tao < dt_den)
    except:
        pass

    danh_sach_day_du = query.order_by(HoSo.ngay_tao.desc()).all()
    loai_ho_so_list = LoaiHoSo.query.order_by(LoaiHoSo.ten_loai).all()

    tong_ho_so = len(danh_sach_day_du)
    tra_lai = sum(1 for hs in danh_sach_day_du if hs.is_tra_lai)
    hoan_tat = sum(1 for hs in danh_sach_day_du if hs.is_hoan_tat() and not hs.is_tra_lai)
    dang_xu_ly = tong_ho_so - tra_lai - hoan_tat

    trang_thai = request.args.get('trang_thai', '')
    danh_sach = danh_sach_day_du
    if trang_thai == 'tra_lai':
        danh_sach = [hs for hs in danh_sach if hs.is_tra_lai]
    elif trang_thai == 'hoan_tat':
        danh_sach = [hs for hs in danh_sach if hs.is_hoan_tat() and not hs.is_tra_lai]
    elif trang_thai == 'dang_xu_ly':
        danh_sach = [hs for hs in danh_sach if not hs.is_hoan_tat() and not hs.is_tra_lai]

    # Tính toán dữ liệu cho Biểu đồ (Loại hồ sơ)
    loai_db_counts = {}
    for hs in danh_sach_day_du:
        loai_ten = hs.loai_ho_so.ten_loai if hs.loai_ho_so else 'Khác'
        loai_db_counts[loai_ten] = loai_db_counts.get(loai_ten, 0) + 1
    
    chart_labels_loai = json.dumps(list(loai_db_counts.keys()))
    chart_data_loai = json.dumps(list(loai_db_counts.values()))

    # Tính toán dữ liệu biểu đồ Trend (Tần suất theo ngày)
    trend_counts = {}
    for hs in sorted(danh_sach_day_du, key=lambda x: x.ngay_tao):
        ds = hs.ngay_tao.strftime('%d/%m')
        trend_counts[ds] = trend_counts.get(ds, 0) + 1
    # Limit to past 14 unique recorded days to not overcrowd the chart
    recent_dates = list(trend_counts.keys())[-14:]
    chart_labels_trend = json.dumps(recent_dates)
    chart_data_trend = json.dumps([trend_counts[k] for k in recent_dates])

    page = request.args.get('page', 1, type=int)
    per_page = 20
    total = len(danh_sach)
    import math
    total_pages = math.ceil(total / per_page) if total > 0 else 1
    danh_sach_paged = danh_sach[(page-1)*per_page : page*per_page]

    # Lấy ghi chú nhanh cho Sidebar
    quick_note = QuickNote.query.first()

    return render_template('index.html', 
        danh_sach=danh_sach_paged, loai_ho_so_list=loai_ho_so_list,
        tu_khoa=tu_khoa, tu_ngay=tu_ngay, den_ngay=den_ngay, loai_id=loai_id, trang_thai=trang_thai,
        dang_xu_ly=dang_xu_ly, hoan_tat=hoan_tat, tra_lai=tra_lai,
        chart_labels_loai=chart_labels_loai, chart_data_loai=chart_data_loai,
        chart_labels_trend=chart_labels_trend, chart_data_trend=chart_data_trend,
        quick_note=quick_note, page=page, total_pages=total_pages
    )

@app.context_processor
def inject_quick_note():
    # Inject quick_note vào mọi template (cần cho Sidebar ở base.html)
    # Thêm check or QuickNote để tránh crash nếu DB trống
    qn = QuickNote.query.first()
    return {'sidebar_quick_note': qn if qn else QuickNote(noi_dung="Chưa có ghi chú...")}

@app.route('/luu_quick_note', methods=['POST'])
@login_required
def luu_quick_note():
    noi_dung = request.form.get('quick_note_content', '')
    qn = QuickNote.query.first()
    if qn:
        qn.noi_dung = noi_dung
        db.session.commit()
        flash("Đã lưu ghi chú sổ tay!", "success")
    return redirect(request.referrer or url_for('index'))

@app.route('/them', methods=['POST'])
@login_required
def add_hoso():
    if request.method == 'POST':
        ma_ho_so = request.form.get('ma_ho_so', '').strip()
        ten_nguoi_dan = request.form.get('ten_nguoi_dan', '').strip()
        loai_id = request.form.get('loai_ho_so_id')
        ghi_chu = request.form.get('ghi_chu', '').strip()
        han_xu_ly_str = request.form.get('han_xu_ly', '').strip()
        
        ngay_sinh = request.form.get('ngay_sinh', '').strip()
        dia_chi = request.form.get('dia_chi', '').strip()
        nhan_khau_kem_theo = request.form.get('nhan_khau_kem_theo', '').strip()
        
        ten_chu_ho = request.form.get('ten_chu_ho', '').strip()
        cccd_chu_ho = request.form.get('cccd_chu_ho', '').strip()
        quan_he_chu_ho = request.form.get('quan_he_chu_ho', '').strip()
        
        if ma_ho_so and ten_nguoi_dan and loai_id:
            loai_id_int = int(loai_id)
            old_hoso = HoSo.query.filter_by(ma_ho_so=ma_ho_so.upper()).first()
            
            if old_hoso:
                flash(f'Mã hồ sơ {ma_ho_so.upper()} đã tồn tại trong hệ thống!', 'danger')
            else:
                buoc_dau_tien = BuocXuLy.query.filter_by(loai_ho_so_id=loai_id_int).order_by(BuocXuLy.so_thu_tu).first()
                
                han_xu_ly_dt = None
                if han_xu_ly_str:
                    try:
                        han_xu_ly_dt = datetime.strptime(han_xu_ly_str, '%Y-%m-%dT%H:%M')
                    except ValueError:
                        try:
                            han_xu_ly_dt = datetime.strptime(han_xu_ly_str, '%Y-%m-%d')
                        except:
                            pass

                new_hoso = HoSo(
                    ma_ho_so=ma_ho_so.upper(),
                    ten_nguoi_dan=ten_nguoi_dan.upper(),
                    ten_khong_dau=tao_chuoi_khong_dau(ten_nguoi_dan),
                    loai_ho_so_id=loai_id_int,
                    buoc_hien_tai_id=buoc_dau_tien.id if buoc_dau_tien else None,
                    ngay_cap_nhat_buoc=datetime.now(),
                    ghi_chu=ghi_chu,
                    han_xu_ly=han_xu_ly_dt,
                    ngay_sinh=ngay_sinh,
                    dia_chi=dia_chi,
                    nhan_khau_kem_theo=nhan_khau_kem_theo,
                    ten_chu_ho=ten_chu_ho.upper() if ten_chu_ho else '',
                    cccd_chu_ho=cccd_chu_ho,
                    quan_he_chu_ho=quan_he_chu_ho
                )
                db.session.add(new_hoso)
                db.session.commit()
                ghi_nhat_ky('Tiếp nhận hồ sơ mới', f'Mã HS: {ma_ho_so.upper()}')
                flash('Thêm hồ sơ mới thành công!', 'success')
        else:
            flash("Vui lòng điền đầy đủ các thông tin bắt buộc!", "warning")
            
    return redirect(url_for('index'))

@app.route('/chuyen_buoc/<int:id>', methods=['POST'])
@login_required
def update_hoso(id):
    hoso = HoSo.query.get_or_404(id)
    if hoso.buoc_hien_tai_id and hoso.buoc_hien_tai:
        next_step = BuocXuLy.query.filter(
            BuocXuLy.loai_ho_so_id == hoso.loai_ho_so_id, 
            BuocXuLy.so_thu_tu > hoso.buoc_hien_tai.so_thu_tu
        ).order_by(BuocXuLy.so_thu_tu).first()
        
        if next_step:
            hoso.buoc_hien_tai_id = next_step.id
            hoso.ngay_cap_nhat_buoc = datetime.now()
            db.session.commit()
            flash(f"Hồ sơ đã chuyển sang bước: {next_step.ten_buoc}", "success")
        else:
            flash("Hồ sơ này đã ở bước cuối cùng!", "info")
    else:
        buoc_dau = BuocXuLy.query.filter_by(loai_ho_so_id=hoso.loai_ho_so_id).order_by(BuocXuLy.so_thu_tu).first()
        if buoc_dau:
            hoso.buoc_hien_tai_id = buoc_dau.id
            hoso.ngay_cap_nhat_buoc = datetime.now()
            db.session.commit()
            flash("Đã thiết lập lại bước xử lý khởi đầu.", "info")
    return redirect(url_for('index'))

@app.route('/tra_lai/<int:id>', methods=['POST'])
@login_required
def tra_lai_hoso(id):
    hoso = HoSo.query.get_or_404(id)
    ly_do = request.form.get('ly_do_tra', '').strip()
    if not hoso.is_tra_lai:
        hoso.is_tra_lai = True
        hoso.ghi_chu = ly_do
        db.session.commit()
        ghi_nhat_ky('Trả lại hồ sơ', f'Mã HS: {hoso.ma_ho_so} - Lý do: {ly_do}')
        flash('Đã đánh dấu hồ sơ là bị TRẢ LẠI!', 'warning')
    return redirect(url_for('index'))

@app.route('/khoi_phuc/<int:id>', methods=['POST'])
@login_required
def khoi_phuc_hoso(id):
    hoso = HoSo.query.get_or_404(id)
    buoc_dau = BuocXuLy.query.filter_by(loai_ho_so_id=hoso.loai_ho_so_id).order_by(BuocXuLy.so_thu_tu).first()
    if buoc_dau:
        hoso.is_tra_lai = False
        hoso.buoc_hien_tai_id = buoc_dau.id
        hoso.ngay_cap_nhat_buoc = datetime.now()
    db.session.commit()
    ghi_nhat_ky('Chuyển bước xử lý', f'Mã HS: {hoso.ma_ho_so}')
    flash('Đã cập nhật trạng thái bước tiếp theo!', 'success')
    return redirect(url_for('index'))

@app.route('/api/cong_dan/<ma_ho_so>')
def api_cong_dan(ma_ho_so):
    hs = HoSo.query.filter_by(ma_ho_so=ma_ho_so.upper()).order_by(HoSo.ngay_tao.desc()).first()
    if hs:
        return jsonify({"found": True, "ten": hs.ten_nguoi_dan})
    return jsonify({"found": False})

@app.route('/api/qr/<int:id>')
def api_qr(id):
    hoso = HoSo.query.get_or_404(id)
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=5, border=2)
    ten_lh = hoso.loai_ho_so.ten_loai if hoso.loai_ho_so else "Khác"
    ma_goc = hoso.ma_ho_so
    ma_formatted = "-".join([ma_goc[i:i+3] for i in range(0, len(ma_goc), 3)]) if ma_goc else ""
    qr_data = f"CCCD: {ma_formatted}\nHo ten: {hoso.ten_nguoi_dan}\nLoai: {ten_lh}\nNgay: {hoso.ngay_tao.strftime('%d/%m/%Y')}"
    qr.add_data(unidecode(qr_data)) 
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img_io = BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    return send_file(img_io, mimetype='image/png')

@app.route('/upload_dinh_kem/<int:id>', methods=['POST'])
@login_required
def upload_dinh_kem(id):
    hoso = HoSo.query.get_or_404(id)
    if 'file' not in request.files:
        flash('Không phát hiện thấy tệp tải lên.', 'danger')
        return redirect(request.referrer or url_for('index'))
    files = request.files.getlist('file')
    count = 0
    for file in files:
        if file.filename == '': continue
        filename = secure_filename(f"{hoso.id}_{datetime.now().strftime('%H%M%S%f')}_{file.filename}")
        file_content = file.read()
        try:
            fp = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            with open(fp, 'wb') as f:
                f.write(file_content)
        except Exception:
            pass
        dk = HoSoDinhKem(ho_so_id=hoso.id, file_name=file.filename, file_path=filename, file_data=file_content)
        db.session.add(dk)
        count += 1
    if count > 0:
        db.session.commit()
        flash(f'Đã lưu thành công {count} tệp đính kèm.', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/xem_file/<path:filename>')
@login_required
def xem_file(filename):
    dk = HoSoDinhKem.query.filter_by(file_path=filename).first()
    if dk and dk.file_data:
        import mimetypes
        mimetype = mimetypes.guess_type(dk.file_name)[0] or 'application/octet-stream'
        return send_file(BytesIO(dk.file_data), mimetype=mimetype, download_name=dk.file_name)
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename))

@app.route('/export_excel')
@login_required
def export_excel():
    tu_khoa = request.args.get('tu_khoa', '').strip()
    loai_id = request.args.get('loai_id', '').strip()
    tu_ngay = request.args.get('tu_ngay', '').strip()
    den_ngay = request.args.get('den_ngay', '').strip()

    query = HoSo.query
    if tu_khoa:
        query = query.filter(or_(HoSo.ma_ho_so.ilike(f'%{tu_khoa}%'), HoSo.ten_nguoi_dan.ilike(f'%{tu_khoa}%')))
    if loai_id and loai_id.isdigit():
        query = query.filter(HoSo.loai_ho_so_id == int(loai_id))
    try:
        if tu_ngay:
            dt_tu = datetime.strptime(tu_ngay, '%Y-%m-%d')
            query = query.filter(HoSo.ngay_tao >= dt_tu)
        if den_ngay:
            from datetime import timedelta
            dt_den = datetime.strptime(den_ngay, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(HoSo.ngay_tao < dt_den)
    except:
        pass

    danh_sach = query.order_by(HoSo.ngay_tao.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo Cáo Hồ Sơ CSKV"
    headers = ['STT', 'CCCD/CMND', 'Họ Tên Công Dân', 'Loại Hồ Sơ', 'Trạng Thái', 'Ngày Tiếp Nhận', 'Lần Xử Lý Cuối', 'Ghi Chú']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 22
    
    for idx, hs in enumerate(danh_sach, 1):
        ngay_cn = hs.ngay_cap_nhat_buoc.strftime('%d/%m/%Y %H:%M') if hs.ngay_cap_nhat_buoc else ''
        row = [
            idx, hs.ma_ho_so, hs.ten_nguoi_dan, 
            hs.loai_ho_so.ten_loai if hs.loai_ho_so else '', 
            hs.get_trang_thai_label(), hs.ngay_tao.strftime('%d/%m/%Y %H:%M'),
            ngay_cn, hs.ghi_chu or ''
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output, as_attachment=True, 
        download_name=f"BaoCao_CSKV_{datetime.now().strftime('%Y%m%d')}.xlsx", 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/backup_db')
@admin_required
def backup_db():
    if os.path.exists(db_path):
        return send_file(db_path, as_attachment=True, download_name=f"Backup_Hoso_{datetime.now().strftime('%Y%m%d_%H%M')}.db")
    else:
        flash("Lỗi hệ thống: Không tìm thấy file Database để sao lưu!", "danger")
        return redirect(request.referrer or url_for('cau_hinh'))

@app.route('/xoa/<int:id>', methods=['POST'])
@login_required
def delete_hoso(id):
    hoso = HoSo.query.get_or_404(id)
    ma_hoso_tam = hoso.ma_ho_so
    ten_hoso_tam = hoso.ten_nguoi_dan
    try:
        for dk in hoso.cac_dinh_kem:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], dk.file_path)
            if os.path.exists(file_path):
                os.remove(file_path)
    except Exception as e:
        print(f"Lỗi khi xóa file vật lý: {e}")
    db.session.delete(hoso)
    db.session.commit()
    flash(f"Đã xoá vĩnh viễn hồ sơ {ma_hoso_tam} ({ten_hoso_tam}) và các tệp đính kèm liên quan.", "info")
    return redirect(url_for('index'))

@app.route('/tra_cuu', methods=['GET'])
@login_required
def tra_cuu():
    tu_khoa = request.args.get('tu_khoa', '').strip()
    ket_qua = []
    
    # Get user quick note for sidebar
    sidebar_quick_note = None
    if current_user:
        tk = TaiKhoan.query.filter_by(username=current_user).first()
        if tk and tk.quick_note:
            sidebar_quick_note = tk.quick_note[0]
            
    if tu_khoa:
        search_pattern = f"%{tu_khoa}%"
        ket_qua = HoSo.query.filter(
            db.or_(
                HoSo.ten_nguoi_dan.ilike(search_pattern),
                HoSo.ma_ho_so.ilike(search_pattern),
                HoSo.ten_chu_ho.ilike(search_pattern),
                HoSo.cccd_chu_ho.ilike(search_pattern),
                HoSo.nhan_khau_kem_theo.ilike(search_pattern)
            )
        ).order_by(HoSo.ngay_tao.desc()).limit(100).all()
        
    return render_template('tra_cuu.html', tu_khoa=tu_khoa, ket_qua=ket_qua, sidebar_quick_note=sidebar_quick_note)

# --- ROUTES CẤU HÌNH ---
@app.route('/cau_hinh')
@admin_required
def cau_hinh():
    loai_ho_so_list = LoaiHoSo.query.all()
    active_loai_id = request.args.get('loai_id', '')
    active_loai = None
    if active_loai_id.isdigit():
        active_loai = LoaiHoSo.query.get(int(active_loai_id))
    elif loai_ho_so_list:
        active_loai = loai_ho_so_list[0]
    return render_template('cau_hinh.html', loai_ho_so_list=loai_ho_so_list, active_loai=active_loai)

@app.route('/them_loai', methods=['POST'])
@admin_required
def them_loai():
    ten_loai = request.form.get('ten_loai', '').strip()
    if ten_loai:
        if LoaiHoSo.query.filter_by(ten_loai=ten_loai).first():
            flash("Tên danh mục này đã có rồi!", "danger")
        else:
            db.session.add(LoaiHoSo(ten_loai=ten_loai))
            db.session.commit()
            flash(f"Đã thêm Loại Hồ sơ mới: {ten_loai}", "success")
    return redirect(url_for('cau_hinh'))

@app.route('/xoa_loai/<int:id>', methods=['POST'])
@admin_required
def xoa_loai(id):
    loai = LoaiHoSo.query.get_or_404(id)
    if loai.ho_so_list:
        flash(f"LỖI: Đang có {len(loai.ho_so_list)} hồ sơ thuộc loại '{loai.ten_loai}' trong hệ thống. Bạn không thể xóa loại này!", "danger")
    else:
        ten_loai_tam = loai.ten_loai
        db.session.delete(loai)
        db.session.commit()
        flash(f"Đã xóa thành công Danh mục {ten_loai_tam}.", "success")
    return redirect(url_for('cau_hinh'))

@app.route('/them_buoc', methods=['POST'])
@admin_required
def them_buoc():
    loai_id = request.form.get('loai_id')
    ten_buoc = request.form.get('ten_buoc', '').strip()
    so_thu_tu = request.form.get('so_thu_tu')
    if loai_id and ten_buoc and so_thu_tu:
        db.session.add(BuocXuLy(loai_ho_so_id=loai_id, so_thu_tu=int(so_thu_tu), ten_buoc=ten_buoc))
        db.session.commit()
        flash("Đã thêm một bước xử lý thành công!", "success")
    else:
        flash("Vui lòng điền đủ tên bước và số thứ tự!", "danger")
    return redirect(url_for('cau_hinh', loai_id=loai_id))

@app.route('/xoa_buoc/<int:buoc_id>', methods=['POST'])
@admin_required
def xoa_buoc(buoc_id):
    buoc = BuocXuLy.query.get_or_404(buoc_id)
    loai_id = buoc.loai_ho_so_id
    hs_dang_o_buoc_nay = HoSo.query.filter_by(buoc_hien_tai_id=buoc_id).count()
    if hs_dang_o_buoc_nay > 0:
        flash(f"Không thể xóa bước này vì đang có {hs_dang_o_buoc_nay} hồ sơ kẹt tại đây. Vui lòng chuyển hồ sơ đi trước!", "danger")
    else:
        db.session.delete(buoc)
        db.session.commit()
        flash("Đã xóa bước xử lý.", "success")
    return redirect(url_for('cau_hinh', loai_id=loai_id))

def open_browser():
    webbrowser.open_new('http://127.0.0.1:5000/')

if __name__ == '__main__':
    if getattr(sys, 'frozen', False):
        Timer(1.2, open_browser).start()
        app.run(host='127.0.0.1', port=5000, debug=False)
    else:
        port = int(os.environ.get('PORT', 5000))
        app.run(host='0.0.0.0', port=port, debug=not os.environ.get('DATABASE_URL'))
